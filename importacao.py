"""
Rotina de importação de dados.

REAL (extraído da planilha oficial `base_upas_feira_de_santana.xlsx`):
  - Cadastro das unidades (aba dim_UPA) → tabela `unidade`
  - Lançamentos financeiros efetivos (aba Fato_Repasses, apenas linhas com
    valor preenchido) → tabela `repasse_financeiro`

ESTIMADO (gerado por função parametrizada, nunca hardcoded, sempre com
±15% de variação controlada e `fonte`/`origem_dado` preenchidos):
  - atendimento_mensal, parametro_custeio, orcamento, classificacao_risco

Onde a planilha tiver "A confirmar" / "Não informado", o campo fica NULO —
nunca preenchido com valor inventado nesta etapa de importação de dado real.
"""
import random
import unicodedata
import openpyxl

from data.database import get_session, reset_db
from data.models import (
    Unidade, AtendimentoMensal, ParametroCusteio, Orcamento,
    ClassificacaoRisco, RepasseFinanceiro,
)
from data import config_estimativas as cfg

NAO_INFORMADO_TOKENS = {
    "a confirmar", "não informado", "nao informado", "-", "", None,
}


def _limpo(valor):
    """Retorna None quando o valor da planilha significa 'não informado'."""
    if valor is None:
        return None
    if isinstance(valor, str):
        v = valor.strip()
        if v.lower() in NAO_INFORMADO_TOKENS:
            return None
        return v
    return valor


def _extrai_porte(porte_raw):
    porte = _limpo(porte_raw)
    if porte in ("Porte I", "Porte II", "Porte III"):
        return porte
    return None


def _extrai_bairro(endereco_raw):
    """Heurística simples para extrair o bairro do texto de endereço da planilha."""
    endereco = _limpo(endereco_raw)
    if not endereco:
        return None
    partes = [p.strip() for p in endereco.split(",")]
    # Remove a cidade/UF final (padrão "Feira de Santana/BA")
    partes = [p for p in partes if "feira de santana" not in p.lower()]
    if len(partes) >= 2:
        return partes[-1]
    return partes[0] if partes else None


def _extrai_natureza_gestao(tipo_gestao_raw):
    texto = (_limpo(tipo_gestao_raw) or "").lower()
    if not texto:
        return None
    if "municipal" in texto or "estadual" in texto:
        # Gestão pública via contrato com OSS/consórcio ainda é natureza Público/SUS
        return "Público/SUS"
    if "privad" in texto:
        return "Privado"
    if "filantróp" in texto or "filantrop" in texto:
        return "Filantrópico/SUS"
    return "Público/SUS"


def _extrai_capacidade(leitos_raw, porte):
    """A planilha às vezes traz a capacidade dentro do campo 'Leitos' (texto livre)."""
    leitos = _limpo(leitos_raw)
    if isinstance(leitos, str) and "atendimentos/mês" in leitos.lower():
        # Ex.: "Capacidade: até 4.500 atendimentos/mês (1.700 m²)"
        import re
        m = re.search(r"até\s*([\d.]+)\s*atendimentos/m", leitos, flags=re.IGNORECASE)
        if m:
            return int(m.group(1).replace(".", ""))
    if porte and porte in cfg.CAPACIDADE_TEORICA_POR_PORTE:
        return cfg.CAPACIDADE_TEORICA_POR_PORTE[porte]
    return None


def importar_unidades(session, caminho_planilha):
    """Lê a aba dim_UPA e popula a tabela `unidade` com dado REAL."""
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb["dim_UPA"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r[0] == "id_UPA")
    header = rows[header_idx]
    col = {name: idx for idx, name in enumerate(header)}

    mapa_id_planilha_para_unidade = {}

    for row in rows[header_idx + 1:]:
        if not row or _limpo(row[col["id_UPA"]]) is None:
            continue
        id_upa_planilha = row[col["id_UPA"]]
        nome = _limpo(row[col["Nome_UPA"]])
        cnes_cnpj = _limpo(row[col.get("CNES/CNPJ")])
        tipo_gestao = _limpo(row[col["Tipo_Gestao"]])
        endereco = _limpo(row[col["Endereço"]])
        porte = _extrai_porte(row[col["Porte"]])
        leitos = row[col["Leitos"]]

        unidade = Unidade(
            nome=nome or f"Unidade {id_upa_planilha} (nome a confirmar)",
            tipo="UPA",
            bairro=_extrai_bairro(endereco),
            natureza_gestao=_extrai_natureza_gestao(tipo_gestao),
            cnes=cnes_cnpj,
            funcionamento="24h",
            porte=porte,
            capacidade_teorica_mensal=_extrai_capacidade(leitos, porte),
            nro_profissionais=cfg.PROFISSIONAIS_POR_PORTE.get(porte) if porte else None,
            origem_cadastro="real",
        )
        session.add(unidade)
        session.flush()  # garante id_unidade antes do commit
        mapa_id_planilha_para_unidade[id_upa_planilha] = unidade.id_unidade

    session.commit()
    return mapa_id_planilha_para_unidade


def importar_repasses_reais(session, caminho_planilha, mapa_unidades):
    """Lê a aba Fato_Repasses e importa apenas as linhas com dado efetivamente
    preenchido (linhas REAIS localizadas em fontes públicas), como dado REAL."""
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb["Fato_Repasses"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r[0] == "id_Lancamento")
    header = rows[header_idx]
    col = {name: idx for idx, name in enumerate(header)}

    esfera_por_fonte = {"FR-01": "Federal", "FR-02": "Federal", "FR-03": "Estadual", "FR-04": "Municipal"}

    total_importado = 0
    for row in rows[header_idx + 1:]:
        if not row or row[col["id_Lancamento"]] is None:
            continue
        id_upa = row[col["id_UPA"]]
        if id_upa not in mapa_unidades:
            continue

        valor_previsto = row[col["Valor_Previsto (R$)"]]
        valor_repassado = row[col["Valor_Repassado (R$)"]]
        valor_pago = row[col["Valor_Pago (R$)"]]

        # Só importa como REAL se houver ao menos um valor financeiro preenchido.
        if valor_previsto is None and valor_repassado is None and valor_pago is None:
            continue

        repasse = RepasseFinanceiro(
            id_unidade=mapa_unidades[id_upa],
            ano=row[col["Ano"]],
            mes=row[col["Mes"]],
            esfera=esfera_por_fonte.get(row[col["id_Fonte"]]),
            instrumento=_limpo(row[col["Instrumento"]]),
            valor_previsto=valor_previsto,
            valor_repassado=valor_repassado,
            valor_pago=valor_pago,
            fonte_dado=_limpo(row[col["Fonte_Dado / Observações"]]),
            origem_dado="real",
        )
        session.add(repasse)
        total_importado += 1

    session.commit()
    return total_importado


def _valor_com_variacao(base, rng, variacao=cfg.VARIACAO_PADRAO):
    fator = 1 + rng.uniform(-variacao, variacao)
    return round(base * fator, 2)


def gerar_dados_simulados(session, seed=42):
    """
    Gera atendimento_mensal, parametro_custeio, orcamento e classificacao_risco
    de forma PARAMETRIZADA (não hardcoded) para cada unidade cadastrada,
    sempre com origem_dado='estimado' e `fonte` preenchida com a premissa usada.

    `seed` fixo por padrão para que o MVP seja reprodutível entre execuções;
    pode ser alterado para gerar novas amostras dentro da mesma metodologia.
    """
    rng = random.Random(seed)
    unidades = session.query(Unidade).all()

    for unidade in unidades:
        porte = unidade.porte or "Porte II"  # premissa: porte médio quando não confirmado no CNES
        custo_base = cfg.CUSTO_BASE_POR_PORTE[porte]
        atendimento_base = cfg.ATENDIMENTOS_BASE_POR_PORTE[porte]

        for periodo in cfg.PERIODOS_SIMULADOS:
            # --- atendimento_mensal (estimado) ---
            qtd_atendimentos = int(round(_valor_com_variacao(atendimento_base, rng)))
            session.add(AtendimentoMensal(
                id_unidade=unidade.id_unidade,
                periodo=periodo,
                qtd_atendimentos=qtd_atendimentos,
                origem_dado="estimado",
            ))

            # --- parametro_custeio (estimado) ---
            custos = {k: _valor_com_variacao(v, rng) for k, v in custo_base.items()}
            session.add(ParametroCusteio(
                id_unidade=unidade.id_unidade,
                periodo=periodo,
                custo_pessoal=custos["custo_pessoal"],
                custo_medicamento=custos["custo_medicamento"],
                custo_material=custos["custo_material"],
                custo_manutencao=custos["custo_manutencao"],
                custo_administrativo=custos["custo_administrativo"],
                tipo_custo_pessoal=cfg.TIPO_CUSTO["custo_pessoal"],
                tipo_custo_medicamento=cfg.TIPO_CUSTO["custo_medicamento"],
                tipo_custo_material=cfg.TIPO_CUSTO["custo_material"],
                tipo_custo_manutencao=cfg.TIPO_CUSTO["custo_manutencao"],
                tipo_custo_administrativo=cfg.TIPO_CUSTO["custo_administrativo"],
                fonte=f"{cfg.FONTE_CUSTEIO_PADRAO} Porte considerado: {porte}.",
            ))

            # --- orcamento (estimado, = soma dos custos estimados do período) ---
            custo_total = sum(custos.values())
            session.add(Orcamento(
                id_unidade=unidade.id_unidade,
                periodo=periodo,
                valor_orcado=round(custo_total, 2),
                origem_dado="estimado",
            ))

            # --- classificacao_risco (estimado) ---
            for categoria, proporcao in cfg.DISTRIBUICAO_RISCO.items():
                session.add(ClassificacaoRisco(
                    id_unidade=unidade.id_unidade,
                    periodo=periodo,
                    categoria_risco=categoria,
                    qtd_estimada=int(round(qtd_atendimentos * proporcao)),
                    origem_dado="estimado",
                ))

    session.commit()


def executar_importacao_completa(caminho_planilha, seed=42):
    """Reseta o banco e executa a importação completa (real + estimado)."""
    reset_db()
    session = get_session()
    try:
        mapa_unidades = importar_unidades(session, caminho_planilha)
        total_repasses = importar_repasses_reais(session, caminho_planilha, mapa_unidades)
        gerar_dados_simulados(session, seed=seed)
        return {
            "unidades_importadas": len(mapa_unidades),
            "repasses_reais_importados": total_repasses,
            "periodos_simulados": len(cfg.PERIODOS_SIMULADOS),
        }
    finally:
        session.close()


if __name__ == "__main__":
    import sys
    caminho = sys.argv[1] if len(sys.argv) > 1 else "base_upas_feira_de_santana.xlsx"
    resultado = executar_importacao_completa(caminho)
    print("Importação concluída:", resultado)
